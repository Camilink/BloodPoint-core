from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from bloodpoint_app.models import campana, donacion

def exportar_top3_campanas_por_donaciones():
    campañas_data = []

    for c in campana.objects.all():
        donaciones = donacion.objects.filter(campana_relacionada=c)
        total_donaciones = donaciones.count()
        total_validadas = donaciones.filter(validada=True).count()
        total_intencionadas = donaciones.filter(es_intencion=True).count()
        total_sangre = sum([d.cantidad_donacion for d in donaciones])
        nuevos_donantes = sum([1 for d in donaciones if d.id_donante.nuevo_donante])

        porcentaje_nuevos = (nuevos_donantes / total_donaciones * 100) if total_donaciones else 0

        try:
            meta_personas = int(c.meta) if c.meta else 0  # Meta como número de personas
        except (ValueError, TypeError):
            meta_personas = 0

        porcentaje_meta = (total_donaciones / meta_personas * 100) if meta_personas else 0

        campañas_data.append({
            "campana": c,
            "total_donaciones": total_donaciones,
            "validadas": total_validadas,
            "intencionadas": total_intencionadas,
            "total_sangre": total_sangre,
            "meta": meta_personas,
            "porc_meta": porcentaje_meta,
            "porc_nuevos": porcentaje_nuevos
        })

    # Ordenar y seleccionar top 3
    top3 = sorted(campañas_data, key=lambda x: x["total_donaciones"], reverse=True)[:3]

    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 3 campañas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "ID campaña", "Nombre", "Fecha inicio", "Fecha fin", "Comuna", "Centro",
        "Representante", "Meta (personas)", "Donaciones recibidas", "% meta alcanzada",
        "Validas", "Intencionadas", "Total sangre (ml)", "% nuevos donantes"
    ]
    ws.append(headers)

    # Aplicar estilo a cabecera
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Agregar datos
    for data in top3:
        c = data["campana"]
        fila = [
            c.id_campana,
            c.nombre_campana,
            c.fecha_campana,
            c.fecha_termino,
            c.id_centro.comuna if c.id_centro else "",
            c.id_centro.nombre_centro if c.id_centro else "",
            f"{c.id_representante.nombre} {c.id_representante.apellido}" if c.id_representante else "",
            data["meta"],
            data["total_donaciones"],
            round(data["porc_meta"], 2),
            data["validadas"],
            data["intencionadas"],
            data["total_sangre"],
            round(data["porc_nuevos"], 2)
        ]
        ws.append(fila)

    # Aplicar estilo a celdas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar ancho de columnas automáticamente
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = length + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    # Guardar en memoria y devolver bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
