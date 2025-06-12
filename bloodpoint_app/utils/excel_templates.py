from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from bloodpoint_app.models import campana
from django.db.models import Sum

def generar_excel_campana(campana_id, response):
    campana_obj = campana.objects.get(id_campana=campana_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Campaña"

    # Estilos - rojo
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # rojo oscuro
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="C00000"),
        right=Side(style="thin", color="C00000"),
        top=Side(style="thin", color="C00000"),
        bottom=Side(style="thin", color="C00000")
    )

    # Cabecera tabla resumen
    headers = [
        "ID Campaña",
        "Nombre Campaña",
        "Fecha Inicio",
        "Fecha Término",
        "Meta Donaciones (unidades)",
        "Donaciones Realizadas (unidades)",
        "Porcentaje Cumplimiento (%)",
        "Total ML Donados",
        "Donantes Únicos",
        "Estado Campaña",
        "Representante Responsable",
    ]
    ws.append(headers)

    # Aplicar estilos a la cabecera
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Datos resumen
    total_donaciones = campana_obj.donacion_set.count()
    total_ml = sum(d.cantidad_donacion for d in campana_obj.donacion_set.all())
    donantes_unicos = campana_obj.donacion_set.values('id_donante').distinct().count()

    meta = int(campana_obj.meta) if campana_obj.meta else 0
    porcentaje = (total_donaciones / meta) * 100 if meta else 0

    representante = str(campana_obj.representante_responsable) if campana_obj.representante_responsable else "Sin representante"

    ws.append([
        campana_obj.id_campana,
        campana_obj.nombre_campana,
        campana_obj.fecha_inicio.strftime('%Y-%m-%d') if campana_obj.fecha_inicio else "",
        campana_obj.fecha_termino.strftime('%Y-%m-%d') if campana_obj.fecha_termino else "",
        campana_obj.meta,
        total_donaciones,
        f"{porcentaje:.1f}%",
        total_ml,
        donantes_unicos,
        campana_obj.estado,
        representante,
    ])

    # Aplicar borde a fila de datos
    for cell in ws[2]:
        cell.border = thin_border
        cell.alignment = center_align

    # Espacio antes de la siguiente tabla
    start_row = 4
    ws.cell(row=start_row, column=1, value="Total ML Donados por Tipo de Sangre")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)

    # Encabezados tipo sangre
    tipo_sangre_header_row = start_row + 1
    ws.cell(row=tipo_sangre_header_row, column=1, value="Tipo de Sangre").font = header_font
    ws.cell(row=tipo_sangre_header_row, column=1).fill = header_fill
    ws.cell(row=tipo_sangre_header_row, column=1).alignment = center_align
    ws.cell(row=tipo_sangre_header_row, column=1).border = thin_border

    ws.cell(row=tipo_sangre_header_row, column=2, value="Total ML Donados").font = header_font
    ws.cell(row=tipo_sangre_header_row, column=2).fill = header_fill
    ws.cell(row=tipo_sangre_header_row, column=2).alignment = center_align
    ws.cell(row=tipo_sangre_header_row, column=2).border = thin_border

    # Tipos de sangre posibles
    tipos_sangre = ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']

    # Donaciones agrupadas por tipo sangre
    donaciones_por_tipo = (
        campana_obj.donacion_set
        .values('id_donante__tipo_sangre')
        .annotate(total_ml=Sum('cantidad_donacion'))
    )
    # Convertir queryset a dict {tipo_sangre: total_ml}
    donaciones_dict = {item['id_donante__tipo_sangre']: item['total_ml'] for item in donaciones_por_tipo}

    current_row = tipo_sangre_header_row + 1
    for tipo in tipos_sangre:
        total_ml_tipo = donaciones_dict.get(tipo, 0)
        ws.cell(row=current_row, column=1, value=tipo)
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border

        ws.cell(row=current_row, column=2, value=total_ml_tipo)
        ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=2).border = thin_border

        current_row += 1

    # Ajustar ancho columnas para mejorar lectura
    column_widths = {
        1: 20,
        2: 25,
        3: 18,
        4: 18,
        5: 25,
        6: 22,
        7: 22,
        8: 18,
        9: 18,
        10: 18,
        11: 30,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(response)
